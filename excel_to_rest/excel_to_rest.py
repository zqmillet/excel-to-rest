from openpyxl import load_workbook

from unicodedata import east_asian_width
from unicodedata import normalize

width_map = {
    'F': 1,     # full width
    'H': 1,     # half width
    'W': 2,     # wide
    'Na': 1,    # narrow
    'A': 1,     # ambiguous
    'N': 1,     # neutral
}

class Canvas:
    def __init__(self):
        self.buffer = ''

    def draw(self, string, end='\n'):
        self.buffer += string + end

def get_char_display_width(char):
    return width_map.get(east_asian_width(char), 1)

def get_string_display_width(string):
    return sum(get_char_display_width(char) for char in normalize('NFC', string))

def _is_in_row_span(row_index, column_index, rowspan):
    rowspan_value = 0
    _row_index = 0

    for index in range(row_index):
        if not (index, column_index) in rowspan.keys():
            continue

        rowspan_value = rowspan[(index, column_index)]
        _row_index = index

    return rowspan_value - (row_index - _row_index) > 0

def _write_cell(canvas, table, y, x, length, rowspan = {}):
    text = table[y][x]
    extra_spaces = ""
    if _is_in_row_span(y, x, rowspan):
        text = "|"
        for i in range(length): #according to column width
            text += " "
        canvas.draw(text, end = "")
    else:
        for i in range(length - get_string_display_width(text) - 2):
            extra_spaces += " " #according to column width
        canvas.draw(f"| {text} " + extra_spaces, end = "")

def _write_column_span_cell(canvas, length, colspan_value): #length argument refers to sum of column widths
    text = ""
    for i in range(length + colspan_value - 1):
        text += " "
    canvas.draw(text, end = "")

def _get_maximum_column_width(table, index): #find the longest cell in the column to set the column's width
    return max(get_string_display_width(row[index]) + 2 for row in table if len(row) > index)

def _get_maximum_row_width(table): #find longest row list (in terms of elements)
    return max(len(row) for row in table)

def _get_total_column_length(table): #collect in a list the widths of each column
    widths = [_get_maximum_column_width(table, i) for i in range(_get_maximum_row_width(table))]
    return widths

def _get_maximum_row_display_width(table): #set the width of the table
    maxi = 0
    for i in range(len(table)):
        cur_len = sum(_get_total_column_length(table)) + len(_get_total_column_length(table)) + 1 # "|" at borders and between cells
        if maxi < cur_len:
            maxi = cur_len
    return maxi

def _draw_border(canvas, table, y, rowspan=None):
    rowspan = rowspan or {}
    col_widths = _get_total_column_length(table)
    length = _get_maximum_row_display_width(table)
    cell_w_count = 0
    cell_counter = 0
    for i in range(length):
        if _is_in_row_span(y, cell_counter - 1, rowspan) and not (i == cell_w_count or i == length - 1):
            canvas.draw(" ", end = "")
        elif i == cell_w_count or i == length - 1:
            canvas.draw("+", end = "")
            if cell_counter != _get_maximum_row_width(table):
                cell_w_count += col_widths[cell_counter] + 1
                cell_counter += 1
        else:
            canvas.draw("-", end = "")
    canvas.draw('')

def tabulate(table, colspan=None, rowspan=None):
    colspan = colspan or {}
    rowspan = rowspan or {}
    canvas = Canvas()

    table = [[str(cell) for cell in row] for row in table]
    table_width = _get_maximum_row_display_width(table)
    col_widths = _get_total_column_length(table)

    for y, row in enumerate(table):
        _draw_border(canvas, table, y, rowspan)
        x = 0
        while x < len(row): #altered for loop
            _write_cell(canvas, table, y, x, col_widths[x], rowspan)
            if (y, x) in colspan.keys():
                colspan_value = colspan[(y, x)]
                _write_column_span_cell(canvas, sum(col_widths[x+1:x+colspan_value]), colspan_value)
                x += colspan_value - 1
            x += 1
        canvas.draw('|')
    _draw_border(canvas, table, _get_maximum_row_width(table) - 1) #close bottom of table
    return canvas.buffer

def excel_to_rest(file_path: str, sheet_name: str) -> str:
    workbook = load_workbook(file_path)
    worksheet = workbook[sheet_name]

    table = [[item.value for item in row] for row in worksheet.rows]

    colspan = {}
    rowspan = {}
    for merged_cell_range in worksheet.merged_cells.ranges:
        coordinate = (merged_cell_range.min_row - 1, merged_cell_range.min_col - 1)
        if merged_cell_range.max_col > merged_cell_range.min_col:
            colspan[coordinate] = merged_cell_range.max_col - merged_cell_range.min_col + 1
        if merged_cell_range.max_row > merged_cell_range.min_row:
            rowspan[coordinate] = merged_cell_range.max_row - merged_cell_range.min_row + 1

    return tabulate(table, colspan, rowspan)
